import tkinter as tk
from tkinter import ttk, messagebox
import threading, time


import numpy as np
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk

# ---- import your original code (no edits) ----
try:
    from ec_sequences import EC_SequenceBuilder, TimeSeriesPrediction
except Exception as e:
    EC_SequenceBuilder = None
    TimeSeriesPrediction = None
    _import_error = e
else:
    _import_error = None

from tkinter import filedialog  # <-- add filedialog

try:
    import pandas as pd
except Exception:
    pd = None

from keithley_controller import KeithleyController
from live_plot_service import LivePlotService
# --- optional: PyVISA just for resource scanning in the GUI ---
try:
    import pyvisa
    _pyvisa_avail = True
except Exception:
    pyvisa = None
    _pyvisa_avail = False
# ---------- Little plotting host on right ----------
class PlotHost:
    def __init__(self, parent):
        self.fig = Figure(figsize=(6, 4), dpi=100)
        self.ax = self.fig.add_subplot(111)
        self.ax.set_title("Ready")
        self.ax.set_xlabel("Sample Index")
        self.ax.set_ylabel("Voltage (V)")
        self.ax.grid(True)

        self.canvas = FigureCanvasTkAgg(self.fig, parent)
        self.canvas_widget = self.canvas.get_tk_widget()
        self.canvas_widget.pack(fill=tk.BOTH, expand=True)

        toolbar_frame = tk.Frame(parent)
        toolbar_frame.pack(fill=tk.X)
        self.toolbar = NavigationToolbar2Tk(self.canvas, toolbar_frame)
        self.toolbar.update()

    def plot(self, y, title="Sequence"):
        self.ax.cla()
        self.ax.grid(True)
        self.ax.set_title(title)
        self.ax.set_xlabel("Sample Index")
        self.ax.set_ylabel("Voltage (V)")
        self.ax.plot(y, marker=".", linewidth=1)
        self.canvas.draw_idle()

class LivePlotWindow(tk.Toplevel):
    """A separate window that holds the live I vs index plot."""
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Live Current Plot")
        self.geometry("800x480")

        self.fig = Figure(figsize=(7.5, 4.2), dpi=100)
        self.ax = self.fig.add_subplot(111)
        self.ax.set_title("Live: I vs Sample Index")
        self.ax.set_xlabel("Sample Index")
        self.ax.set_ylabel("Current (A)")
        self.ax.grid(True)

        self.canvas = FigureCanvasTkAgg(self.fig, self)
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        tb_frame = tk.Frame(self)
        tb_frame.pack(fill=tk.X)
        NavigationToolbar2Tk(self.canvas, tb_frame).update()

# ---------- Simple dynamic form helper ----------
class ParamForm(ttk.Frame):
    def __init__(self, parent, fields):
        super().__init__(parent)
        self.vars = {}
        for i, (name, default) in enumerate(fields):
            row = ttk.Frame(self)
            row.pack(fill=tk.X, pady=3)
            ttk.Label(row, text=name + ":", width=18).pack(side=tk.LEFT)

            # --- NEW: dropdown if default is a sequence (list/tuple) ---
            if isinstance(default, (list, tuple)) and len(default) > 0:
                options = list(default)
                var = tk.StringVar(value=str(options[0]))
                self.vars[name] = var
                cb = ttk.Combobox(row, textvariable=var, state="readonly",
                                  values=[str(x) for x in options])
                cb.pack(side=tk.LEFT, fill=tk.X, expand=True)
            else:
                # old behavior: plain text field
                var = tk.StringVar(value=str(default))
                self.vars[name] = var
                ttk.Entry(row, textvariable=var).pack(side=tk.LEFT, fill=tk.X, expand=True)

    def values(self):
        return {k: v.get().strip() for k, v in self.vars.items()}



# ---------- The main App (only calls your builder functions) ----------
# ---------- The main App (only calls your builder functions) ----------
class App:
    def __init__(self, root):
        self.root = root
        self.builder = EC_SequenceBuilder() if EC_SequenceBuilder else None
        self.device = KeithleyController()
        self.last_y = None
        self.last_data = []
        self.output_thread = None

        # Window & layout
        root.title("EC Sequences — GUI")
        root.geometry("1200x800")

        main = tk.Frame(root)
        main.pack(fill=tk.BOTH, expand=True)

        left = tk.Frame(main, padx=10, pady=10, bg="#eee")
        left.pack(side=tk.LEFT, fill=tk.Y)

        right = tk.Frame(main, padx=10, pady=10)
        right.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        # Right plot (voltage / summary plots)
        self.plot_host = PlotHost(right)

        # Live-plot service (can target panel or a separate window)
        self.live = LivePlotService(root, self.plot_host.ax, self.plot_host.canvas,
                            maxlen=5000, max_fps=20, manage_axis_text=False)


        self.live.start()
        self.live_window = None  # created on demand

        # Sequence selector
        tk.Label(left, text="Select generator:", font=("Segoe UI", 11, "bold"), bg="#eee").pack(anchor="w")

        self.combo = ttk.Combobox(
            left,
            state="readonly",
            values=[
                "generate_triangle_wave",
                "generate_sine_wave",
                "generate_ppf_sequence",
                "generate_srdp_sequence",
                "generate_ltp_ltd_sequence",
                "generate_read_sequence",
                "generate_retention_sequence",
                "generate_square_pulse",
                "generate_triangle_pulse",
                "build_stdp_sequence",
                "generate_tstdp_sequence",
                "load_sequence_from_file",
            ],
        )
        self.combo.current(0)
        self.combo.pack(fill=tk.X, pady=6)
        self.combo.bind("<<ComboboxSelected>>", lambda _e: self.load_form())

        # Params panel host
        self.param_panel = ttk.Frame(left)
        self.param_panel.pack(fill=tk.X, pady=(4, 10))

        # Generate button
        ttk.Button(left, text="Generate & Plot", command=self.run_current).pack(fill=tk.X, pady=4)

        # --- Device / Output panel ---
        ttk.Separator(left).pack(fill=tk.X, pady=8)
        ttk.Label(left, text="Instrument Control", font=("Segoe UI", 11, "bold"), background="#eee").pack(anchor="w", pady=(0, 4))

        dev_frame = ttk.Frame(left)
        dev_frame.pack(fill=tk.X, pady=2)

        # row 0: VISA address + refresh
        ttk.Label(dev_frame, text="VISA Address:").grid(row=0, column=0, sticky="w")
        self.addr_var = tk.StringVar()
        self.addr_combo = ttk.Combobox(dev_frame, textvariable=self.addr_var, state="readonly", values=[])
        self.addr_combo.grid(row=0, column=1, sticky="ew", padx=4)
        ttk.Button(dev_frame, text="Refresh", command=self.scan_visa).grid(row=0, column=2, sticky="w")
        dev_frame.columnconfigure(1, weight=1)

        # row 1: compliance
        ttk.Label(dev_frame, text="Compliance (A):").grid(row=1, column=0, sticky="w", pady=(4, 0))
        self.ilim_var = tk.StringVar(value="0.01")
        ttk.Entry(dev_frame, textvariable=self.ilim_var, width=12).grid(row=1, column=1, sticky="w", padx=4, pady=(4, 0))

        # row 2: dwell
        ttk.Label(dev_frame, text="Dwell per step (ms):").grid(row=2, column=0, sticky="w")
        self.dwell_var = tk.StringVar(value="10")
        ttk.Entry(dev_frame, textvariable=self.dwell_var, width=12).grid(row=2, column=1, sticky="w", padx=4)

        # row 3: live plot on/off
        ttk.Label(dev_frame, text="Live Plot:").grid(row=3, column=0, sticky="w")
        self.live_var = tk.StringVar(value="Off")
        ttk.Combobox(dev_frame, textvariable=self.live_var, state="readonly", values=["Off", "On"]).grid(row=3, column=1, sticky="w", padx=4)

        # ADD: 分析图选择
        ttk.Label(dev_frame, text="Analysis plot:").grid(row=4, column=0, sticky="w")
        self.analysis_var = tk.StringVar(value="I vs sample index")
        ttk.Combobox(
            dev_frame, textvariable=self.analysis_var, state="readonly",
            values=["I vs time", "I vs V", "I vs sample index"]
            ).grid(row=4, column=1, sticky="w", padx=4)


        # row 5: channel mode / polarity mapping
        ttk.Label(dev_frame, text="Channel mode:").grid(row=5, column=0, sticky="w")
        self.chan_mode_var = tk.StringVar(value="A only")
        ttk.Combobox(
            dev_frame, textvariable=self.chan_mode_var, state="readonly",
            values=["A only", "B only", "Dual: A = +, B = −", "Dual: A = −, B = +"]
            ).grid(row=5, column=1, sticky="w", padx=4)
        # initial VISA scan
        self.root.after(100, self.scan_visa)

        # Connect / Disconnect buttons
        btn_row = ttk.Frame(left)
        btn_row.pack(fill=tk.X, pady=(6, 2))
        ttk.Button(btn_row, text="Connect", command=self.connect_device).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(0, 4))
        ttk.Button(btn_row, text="Disconnect", command=self.disconnect_device).pack(side=tk.LEFT, expand=True, fill=tk.X)

        # Run / Stop buttons
        run_row = ttk.Frame(left)
        run_row.pack(fill=tk.X, pady=(4, 2))
        ttk.Button(run_row, text="Start Output (last sequence)", command=self.start_output).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(0, 4))
        ttk.Button(run_row, text="Stop", command=self.stop_output).pack(side=tk.LEFT, expand=True, fill=tk.X)

        # Save button
        save_row = ttk.Frame(left)
        save_row.pack(fill=tk.X, pady=(4, 2))
        ttk.Button(save_row, text="Save Log CSV", command=self.save_log_csv).pack(side=tk.LEFT, expand=True, fill=tk.X)

        # Status label
        self.status_var = tk.StringVar(value="Instrument: DISCONNECTED")
        ttk.Label(
            left,
            textvariable=self.status_var,
            background="#eee",
            foreground="#333",
            wraplength=260,
            justify="left",
        ).pack(anchor="w", pady=(6, 0))

        # Info
        ttk.Separator(left).pack(fill=tk.X, pady=8)
        ttk.Label(
            left,
            text="Tip: type numbers; lists as comma-separated\n(e.g. 30,15,5,15)",
            background="#eee",
            foreground="#333",
            justify="left",
        ).pack(anchor="w")

        # Build first form
        self.current_form = None
        self.current_form_name = None
        self.load_form()


    def scan_visa(self):
        """Populate the VISA address combobox. Prefer GPIB resources."""
        resources = []
        if _pyvisa_avail:
            try:
                rm = pyvisa.ResourceManager()
                resources = list(rm.list_resources())
            except Exception as e:
                messagebox.showwarning("VISA scan failed", str(e))
                resources = []

        gpib = [r for r in resources if str(r).upper().startswith("GPIB")]
        others = [r for r in resources if r not in gpib]
        ordered = gpib + others

        if not ordered:
            self.addr_combo["values"] = ["<no VISA resources found>"]
            if not self.addr_var.get():
                self.addr_var.set("<no VISA resources found>")
            return

        self.addr_combo["values"] = ordered
        if self.addr_var.get() not in ordered:
            self.addr_var.set(ordered[0])



    def _close_live_window(self):
        # Safety: abort the run and force outputs to 0/Off
        try:
            try:
                self.device.stop()          # request loop to abort
            except Exception:
                pass
            try:
                self.device.safe_zero_all() # hard safety
            except Exception:
                pass
        finally:
            try:
                if self.live_window and self.live_window.winfo_exists():
                    self.live_window.destroy()
            finally:
                self.live_window = None
                try:
                    self.status_var.set("Live window closed — outputs OFF, level=0 (safe).")
                except Exception:
                    pass

    def _ensure_live_window(self):
        """确保存在独立的 Live/分析窗口，并把 live service 画到它。"""
        if not (self.live_window and self.live_window.winfo_exists()):
            self.live_window = LivePlotWindow(self.root)
            self.live_window.protocol("WM_DELETE_WINDOW", self._close_live_window)

    
    def _config_axes_for_mode(self, ax, xmode, title=None, clear=True):
        if clear:
            ax.cla()
        ax.grid(True)
        ax.set_ylabel("Current (A)")
        if xmode == "time":
            ax.set_xlabel("Time (s)")
            ax.set_title(title or "Live: I vs Time")
        elif xmode == "v":
            ax.set_xlabel("Voltage (V)")
            ax.set_title(title or "Live: I vs V")
        else:
            ax.set_xlabel("Sample Index")
            ax.set_title(title or "Live: I vs Sample Index")



    def _plot_analysis_in_window(self, dwell_s: float):
        """把测量完成后的分析图画到独立窗口里。"""
        if not self.last_data:
            return

        # 取数据
        idx = np.array([d.get("index", i) for i, d in enumerate(self.last_data)], dtype=float)
        # 时间：优先用回调里记录的 t，否则用均匀采样估计
        t = np.array([d.get("t", i * dwell_s) for i, d in enumerate(self.last_data)], dtype=float)
        t = t - (t[0] if len(t) else 0.0)
        v = np.array([d.get("V_meas", d.get("V_set", np.nan)) for d in self.last_data], dtype=float)
        i = np.array([d.get("I_meas", np.nan) for d in self.last_data], dtype=float)

        # 选择横轴：优先用运行时缓存的 _current_xmode
        sel = getattr(self, "_current_xmode", None)
        if not sel:
            sel_txt = (self.analysis_var.get() or "I vs sample index").lower()
            if "time" in sel_txt:
                sel = "time"
            elif " v" in sel_txt or sel_txt.endswith("v"):
                sel = "v"
            else:
                sel = "index"

        if sel == "time":
            x, xlabel, title = t, "Time (s)", "I vs time"
        elif sel == "v":
            x, xlabel, title = v, "Voltage (V)", "I vs V"
        else:
            x, xlabel, title = idx, "Sample Index", "I vs sample index"


        # 画到独立窗口
        self._ensure_live_window()
        self.live_window.deiconify()
        self.live_window.lift()
        self.live_window.focus_force()
        ax = self.live_window.ax
        ax.cla()
        ax.grid(True)
        ax.set_title(title)
        ax.set_xlabel(xlabel)
        ax.set_ylabel("Current (A)")
        ax.plot(x, i, marker=".", linewidth=1, label="I_meas")
        ax.legend()
        self.live_window.canvas.draw_idle()


    # ----- forms for each method (names exactly match your EC_SequenceBuilder) -----
    def load_form(self):
        for w in self.param_panel.winfo_children():
            w.destroy()

        name = self.combo.get()
        self.current_form_name = name
        if name == "generate_triangle_wave":
            fields = [
                ("amplitude", 1.0),
                ("n_points", 100),
                ("n_cycles", 1),
                ("reverse", [False, True])  # dropdown: False or True
            ]
        elif name == "generate_ppf_sequence":
            fields = [("a", 0.8), ("n_pulse", 1), ("n_gap", 10), ("n_rest", 0)]
        elif name == "generate_srdp_sequence":
            fields = [("V_write", 0.8), ("n_pulse", 5), ("gap_list", "30,10,5,10"), ("repeat", 10)]
        elif name == "generate_ltp_ltd_sequence":
            fields = [("V_read", 0.1), ("n_read", 5),
                      ("V_ltp", 0.8), ("n_ltp_pulse", 5),
                      ("V_ltd", -1.0), ("n_ltd_pulse", 5),
                      ("n_block1", 10), ("n_block2", 10)]
        elif name == "generate_read_sequence":
            fields = [
                ("V_read", 0.1),
                ("n_total", 40),
                ("mode", ["const", "discre", "alt"]),  # dropdown
                ("n_alt", 10)
            ]
        elif name == "generate_retention_sequence":
            # read_params_* are dicts; we input minimal keys as simple fields
            fields = [
                ("read_params_pre.V_read", 0.1),
                ("read_params_pre.n_total", 40),
                ("read_params_pre.mode", ["const", "discre", "alt"]),  # dropdown
                ("n_read_pre", 5),

                ("read_params_post.V_read", 0.1),
                ("read_params_post.n_total", 40),
                ("read_params_post.mode", ["const", "discre", "alt"]),  # dropdown
                ("n_read_post", 2),

                ("V_write", 1.0),
                ("n_pulse", 5),
                ("n_gap", 2),
                ("cycle", 10),
            ]
        elif name == "generate_square_pulse":
            fields = [("V", 1.0), ("length", 10), ("space", 5)]
        elif name == "generate_triangle_pulse":
            fields = [("V", 1.0), ("step", 10), ("space", 0)]
        elif name == "build_stdp_sequence":
            fields = [
                # read sequence
                ("read_seq.V", 0.1),
                ("read_seq.n_total", 10),
                ("read_seq.mode", ["const", "alt", "discre"]),  # dropdown
                ("read_seq.n_alt", 10),

                # pulse type dropdowns + parameters (triangle uses 'step', square uses 'length')
                ("pre_pulse_type",  ["triangle", "square"]),
                ("pre_seq.V", 1.0),
                ("pre_seq.step_or_len", 10),
                ("pre_seq.space", 0),

                ("post_pulse_type", ["triangle", "square"]),
                ("post_seq.V", -1.0),
                ("post_seq.step_or_len", 10),
                ("post_seq.space", 0),

                # STDP-specific
                ("delay_points", 100),
                ("read_len", 10),                     # usually = read_seq.n_total
                ("stdp_mode", ["bef", "alt"]),      # dropdown
            ]

        elif name == "generate_tstdp_sequence":
            fields = [
                ("read_seq.V", 0.1),
                ("read_seq.n_total", 10),
                ("read_seq.mode", ["const", "alt", "discre"]),  # dropdown (first item is initial)
                ("read_seq.n_alt", 10),

                # pulse type dropdowns: first item is initial selection
                ("pre_pulse_type", ["triangle", "square"]),
                ("pre_seq.V", 1.0),
                ("pre_seq.step_or_len", 10),
                ("pre_seq.space", 0),

                ("post_pulse_type", ["triangle",  "square"]),
                ("post_seq.V", -1.0),
                ("post_seq.step_or_len", 10),
                ("post_seq.space", 0),

                ("mode", ["post-pre-post", "pre-post-pre", "pre-pre-post",
                          "post-pre-post", "post-post-pre", "pre-post-post"]),  # tSTDP mode dropdown
                ("t_1", -20),
                ("t_2", 100),
            ]

        elif name == "generate_sine_wave":
            fields = [
                ("amplitude", 1.0),
                ("n_points", 100),                    # samples per cycle if not using freq
                ("n_cycles", 1),
                ("start_at", ["zero_rise", "pos_max", "neg_max", "zero_fall", "0deg", "90deg", "180deg", "270deg"]),
                ("end_at",   ["<none>", "zero_rise", "pos_max", "neg_max", "zero_fall", "0deg", "90deg", "180deg", "270deg"]),
                # Optional alternative to n_points:
                ("freq_hz",  ""),                     # leave blank to ignore
                ("dwell_s",  ""),                     # leave blank to ignore
                ("dc_offset", 0.0),
                ("include_endpoint", [False, True]),  # dropdown bool
            ]

        elif name == "load_sequence_from_file":
            fields = [
                ("skip_rows", 1)
            ]  # one input; user will browse for a file

        else:
            fields = []

        self.current_form = ParamForm(self.param_panel, fields)
        self.current_form.pack(fill=tk.X)

    # ----- helpers to parse common field patterns without changing your logic -----
    @staticmethod
    def _to_float(s): return float(s)
    @staticmethod
    def _to_int(s): return int(float(s))
    @staticmethod
    def _to_optional_float(s):
        """Return float(s) or None if blank/'none'."""
        if s is None:
            return None
        t = str(s).strip().lower()
        if t in ("", "none", "<none>"):
            return None
        return float(t)

    @staticmethod
    def _to_bool(s): return str(s).strip().lower() in ("1", "true", "t", "yes", "y")
    @staticmethod
    def _to_list_ints(s): return [int(float(x)) for x in s.replace("，", ",").split(",") if x.strip()]
    @staticmethod
    def _to_list_floats(s): return [float(x) for x in s.replace("，", ",").split(",") if x.strip()]

    # ----- run selected generator and plot -----
    def run_current(self):
        if not self.builder:
            messagebox.showerror("Import error", f"Cannot import ec_sequences.py:\n{_import_error}")
            return

        name = self.combo.get()
        p = self.current_form.values()

        try:
            if name == "generate_triangle_wave":
                y = self.builder.generate_triangle_wave(
                    amplitude=self._to_float(p["amplitude"]),
                    n_points=self._to_int(p["n_points"]),
                    n_cycles=self._to_int(p["n_cycles"]),
                    reverse=self._to_bool(p["reverse"]),
                )
                self.last_yA = np.asarray(y, dtype=float)
                self.last_yB = None
                self.last_y  = self.last_yA  # keep legacy field populated
            elif name == "generate_ppf_sequence":
                y = self.builder.generate_ppf_sequence(
                    a=self._to_float(p["a"]),
                    n_pulse=self._to_int(p["n_pulse"]),
                    n_gap=self._to_int(p["n_gap"]),
                    n_rest=self._to_int(p["n_rest"]),
                )
                self.last_yA = np.asarray(y, dtype=float)
                self.last_yB = None
                self.last_y  = self.last_yA  # keep legacy field populated

            elif name == "generate_srdp_sequence":
                y = self.builder.generate_srdp_sequence(
                    V_write=self._to_float(p["V_write"]),
                    n_pulse=self._to_int(p["n_pulse"]),
                    gap_list=self._to_list_ints(p["gap_list"]),
                    repeat=self._to_int(p["repeat"]),
                )
                self.last_yA = np.asarray(y, dtype=float)
                self.last_yB = None
                self.last_y  = self.last_yA  # keep legacy field populated
            elif name == "generate_ltp_ltd_sequence":
                y = self.builder.generate_ltp_ltd_sequence(
                    V_read=self._to_float(p["V_read"]),
                    n_read=self._to_int(p["n_read"]),
                    V_ltp=self._to_float(p["V_ltp"]),
                    n_ltp_pulse=self._to_int(p["n_ltp_pulse"]),
                    V_ltd=self._to_float(p["V_ltd"]),
                    n_ltd_pulse=self._to_int(p["n_ltd_pulse"]),
                    n_block1=self._to_int(p["n_block1"]),
                    n_block2=self._to_int(p["n_block2"]),
                )
                self.last_yA = np.asarray(y, dtype=float)
                self.last_yB = None
                self.last_y  = self.last_yA  # keep legacy field populated
            elif name == "generate_read_sequence":
                y = self.builder.generate_read_sequence(
                    V_read=self._to_float(p["V_read"]),
                    n_total=self._to_int(p["n_total"]),
                    mode=p["mode"],
                    n_alt=self._to_int(p["n_alt"]),
                )
                self.last_yA = np.asarray(y, dtype=float)
                self.last_yB = None
                self.last_y  = self.last_yA  # keep legacy field populated
            elif name == "generate_retention_sequence":
                read_params_pre = dict(V_read=self._to_float(p["read_params_pre.V_read"]),
                                       n_total=self._to_int(p["read_params_pre.n_total"]),
                                       mode=p["read_params_pre.mode"])
                read_params_post = dict(V_read=self._to_float(p["read_params_post.V_read"]),
                                        n_total=self._to_int(p["read_params_post.n_total"]),
                                        mode=p["read_params_post.mode"])
                y = self.builder.generate_retention_sequence(
                    read_params_pre=read_params_pre, n_read_pre=self._to_int(p["n_read_pre"]),
                    read_params_post=read_params_post, n_read_post=self._to_int(p["n_read_post"]),
                    V_write=self._to_float(p["V_write"]), n_pulse=self._to_int(p["n_pulse"]),
                    n_gap=self._to_int(p["n_gap"]), cycle=self._to_int(p["cycle"]),
                )
                self.last_yA = np.asarray(y, dtype=float)
                self.last_yB = None
                self.last_y  = self.last_yA  # keep legacy field populated
            elif name == "generate_square_pulse":
                y = self.builder.generate_square_pulse(
                    V=self._to_float(p["V"]),
                    length=self._to_int(p["length"]),
                    space=self._to_int(p["space"]),
                )
                self.last_yA = np.asarray(y, dtype=float)
                self.last_yB = None
                self.last_y  = self.last_yA  # keep legacy field populated
            elif name == "generate_triangle_pulse":
                y = self.builder.generate_triangle_pulse(
                    V=self._to_float(p["V"]),
                    step=self._to_int(p["step"]),
                    space=self._to_int(p["space"]),
                )
                self.last_yA = np.asarray(y, dtype=float)
                self.last_yB = None
                self.last_y  = self.last_yA  # keep legacy field populated
            elif name == "build_stdp_sequence":
                # 1) read sequence ...
                read_seq = self.builder.generate_read_sequence(
                    V_read=self._to_float(p["read_seq.V"]),
                    n_total=self._to_int(p["read_seq.n_total"]),
                    mode=p["read_seq.mode"],
                    n_alt=self._to_int(p["read_seq.n_alt"]),
                )

                def make_pulse(pulse_type, V, step_or_len, space):
                    pt = str(pulse_type).strip().lower()
                    if pt == "triangle":
                        return self.builder.generate_triangle_pulse(
                            V=self._to_float(V),
                            step=self._to_int(step_or_len),
                            space=self._to_int(space),
                        )
                    elif pt == "square":
                        return self.builder.generate_square_pulse(
                            V=self._to_float(V),
                            length=self._to_int(step_or_len),
                            space=self._to_int(space),
                        )
                    else:
                        raise ValueError("pulse_type must be 'triangle' or 'square'")

                spike_pre  = make_pulse(p["pre_pulse_type"],  p["pre_seq.V"],  p["pre_seq.step_or_len"],  p["pre_seq.space"])
                spike_post = make_pulse(p["pre_pulse_type"],  p["pre_seq.V"],  p["pre_seq.step_or_len"],  p["pre_seq.space"])
                reset_post = make_pulse(p["post_pulse_type"], p["post_seq.V"], p["post_seq.step_or_len"], p["post_seq.space"])

                pre_seq  = np.concatenate([spike_pre,  reset_post])
                post_seq = np.concatenate([spike_post, reset_post])

                vseq_a, vseq_b = self.builder.build_stdp_sequence(
                    read_seq=read_seq,
                    pre_seq=pre_seq,
                    post_seq=post_seq,
                    delay_points=self._to_int(p["delay_points"]),
                    read_len=self._to_int(p["read_len"]),
                    mode=p["stdp_mode"],
                )

                # Plot both
                self.plot_host.ax.cla()
                self.plot_host.ax.grid(True)
                self.plot_host.ax.set_title("build_stdp_sequence")
                self.plot_host.ax.set_xlabel("Sample Index")
                self.plot_host.ax.set_ylabel("Voltage (V)")
                self.plot_host.ax.plot(vseq_a, marker=".", linewidth=1, label="A")
                self.plot_host.ax.plot(vseq_b, marker="o", linewidth=1, label="B")
                self.plot_host.ax.legend()
                self.plot_host.canvas.draw_idle()
                
                # 正确的缓存：同时保存 A/B 以及兼容的 last_y
                self.last_yA = np.asarray(vseq_a, dtype=float)
                self.last_yB = np.asarray(vseq_b, dtype=float)
                self.last_y  = self.last_yA
                return


            elif name == "generate_tstdp_sequence":
                # 1) read sequence
                read_seq = self.builder.generate_read_sequence(
                    V_read=self._to_float(p["read_seq.V"]),
                    n_total=self._to_int(p["read_seq.n_total"]),
                    mode=p["read_seq.mode"],
                    n_alt=self._to_int(p["read_seq.n_alt"]),
                )

                # helper: choose triangle vs square using the same "step_or_len" input
                def make_pulse(pulse_type, V, step_or_len, space):
                    if str(pulse_type).strip().lower() == "triangle":
                        return self.builder.generate_triangle_pulse(
                            V=self._to_float(V),
                            step=self._to_int(step_or_len),
                            space=self._to_int(space),
                        )
                    elif str(pulse_type).strip().lower() == "square":
                        return self.builder.generate_square_pulse(
                            V=self._to_float(V),
                            length=self._to_int(step_or_len),  # length for square
                            space=self._to_int(space),
                        )
                    else:
                        raise ValueError("pulse_type must be 'triangle' or 'square'.")

                # 2) pulses
                spike_pre  = make_pulse(p["pre_pulse_type"],  p["pre_seq.V"],  p["pre_seq.step_or_len"],  p["pre_seq.space"])
                reset_post = make_pulse(p["post_pulse_type"], p["post_seq.V"], p["post_seq.step_or_len"], p["post_seq.space"])
                spike_post = make_pulse(p["pre_pulse_type"],  p["pre_seq.V"],  p["pre_seq.step_or_len"],  p["pre_seq.space"])
                # concatenate like in your ec-sequences example
                pre_seq  = np.concatenate([spike_pre,  reset_post])
                post_seq = np.concatenate([spike_post, reset_post])

                # 3) call your original function
                y_a, y_b = self.builder.generate_tstdp_sequence(
                    read_seq=read_seq,
                    pre_seq=pre_seq,
                    post_seq=post_seq,
                    mode=p["mode"],
                    t_1=self._to_int(p["t_1"]),
                    t_2=self._to_int(p["t_2"]),
                )

                # 4) plot both channels
                self.plot_host.ax.cla()
                self.plot_host.ax.grid(True)
                self.plot_host.ax.set_title("generate_tstdp_sequence")
                self.plot_host.ax.set_xlabel("Sample Index")
                self.plot_host.ax.set_ylabel("Voltage (V)")
                self.plot_host.ax.plot(y_a, marker=".", linewidth=1, label="A")
                self.plot_host.ax.plot(y_b, marker="o", linewidth=1, label="B")
                self.plot_host.ax.legend()
                self.plot_host.canvas.draw_idle()

                # 5) cache for channel output
                try:
                    self.last_yA = np.asarray(y_a, dtype=float)
                    self.last_yB = np.asarray(y_b, dtype=float)
                    self.last_y  = self.last_yA
                except Exception:
                    self.last_yA, self.last_yB = None, None
                    self.last_y = None

                return  # IMPORTANT: skip the generic single-series plot below

            elif name == "generate_sine_wave":
                # Optional frequency path
                freq_hz  = self._to_optional_float(p.get("freq_hz", ""))
                dwell_s_ = self._to_optional_float(p.get("dwell_s", ""))

                # If using frequency+dwell, n_points must be None
                n_points_arg = None if freq_hz is not None else self._to_int(p["n_points"])

                # end_at can be "<none>" -> None
                end_at = None if str(p.get("end_at", "")).strip().lower() in ("", "<none>", "none") else p["end_at"]

                y = self.builder.generate_sine_wave(
                    amplitude=self._to_float(p["amplitude"]),
                    n_points=n_points_arg,
                    n_cycles=self._to_int(p["n_cycles"]),
                    start_at=p["start_at"],
                    end_at=end_at,
                    freq_hz=freq_hz,
                    dwell_s=dwell_s_,
                    dc_offset=self._to_float(p["dc_offset"]),
                    include_endpoint=self._to_bool(p["include_endpoint"]),
                )

                # cache & plot
                self.last_yA = np.asarray(y, dtype=float)
                self.last_yB = None
                self.last_y  = self.last_yA
                self.plot_host.plot(y, title="generate_sine_wave")
                return

            elif name == "load_sequence_from_file":
                path = filedialog.askopenfilename(
                    title="Select voltage file",
                    filetypes=[
                        ("All supported", "*.csv *.txt *.xlsx *.xls"),
                        ("CSV", "*.csv"),
                        ("Text", "*.txt"),
                        ("Excel", "*.xlsx *.xls"),
                        ("All files", "*.*"),
                    ]
                )
                if not path:
                    return  # user canceled

                skip_rows = self._to_int(p.get("skip_rows", 0))
                y = None

                try:
                    lower = path.lower()
                    if lower.endswith(".csv") or lower.endswith(".txt"):
                        try:
                            data = np.loadtxt(path, delimiter=",", dtype=float, skiprows=skip_rows)
                        except Exception:
                            data = np.loadtxt(path, dtype=float, skiprows=skip_rows)
                        data = np.atleast_2d(data)
                        y = data[:, 0]

                    elif lower.endswith(".xlsx") or lower.endswith(".xls"):
                        if pd is not None:
                            df = pd.read_excel(path, header=None, skiprows=skip_rows)
                            y = df.iloc[:, 0].dropna().astype(float).to_numpy()
                        else:
                            from openpyxl import load_workbook
                            wb = load_workbook(path, read_only=True, data_only=True)
                            ws = wb.active
                            col_vals = []
                            for idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True)):
                                if idx < skip_rows:
                                    continue
                                v = row[0]
                                if v is None:
                                    continue
                                try:
                                    col_vals.append(float(v))
                                except Exception:
                                    pass
                            y = np.array(col_vals, dtype=float)

                    else:
                        raise ValueError("Unsupported file type.")

                    if y is None or len(y) == 0:
                        raise ValueError("No numeric values found in the first column.")

                    self.plot_host.ax.cla()
                    self.plot_host.ax.grid(True)
                    self.plot_host.ax.set_title("Loaded sequence from file")
                    self.plot_host.ax.set_xlabel("Sample Index")
                    self.plot_host.ax.set_ylabel("Voltage (V)")
                    self.plot_host.ax.plot(y, marker=".", linewidth=1, label="Voltage")
                    self.plot_host.ax.legend()
                    self.plot_host.canvas.draw_idle()

                except Exception as e:
                    messagebox.showerror("Load error", str(e))



            else:
                messagebox.showwarning("Unknown", f"Unknown generator: {name}")
                return

            # Plot
            title = f"{name}"
            self.plot_host.plot(y, title=title)
            try:
                self.last_y = np.asarray(y, dtype=float)
            except Exception:
                self.last_y = None

        except Exception as e:
            messagebox.showerror("Error", str(e))
    
    # ===== Instrument control methods =====
    def connect_device(self):
        addr = self.addr_var.get().strip()
        try:
            msg = self.device.connect(addr)
            self.status_var.set(f"Instrument: CONNECTED — {msg}")
        except Exception as e:
            messagebox.showerror("Connect error", str(e))
            self.status_var.set("Instrument: DISCONNECTED")

    def disconnect_device(self):
        try:
            self.device.disconnect()
            self.status_var.set("Instrument: DISCONNECTED")
        except Exception as e:
            messagebox.showerror("Disconnect error", str(e))

    def start_output(self):
        """Stream the last generated sequence to the instrument with channel selection,
        optional dual-channel polarity mapping, live plotting to a separate window,
        and end-of-run analysis plotted only in that window.
        """
        # 0) Ensure we have a sequence
        if getattr(self, "last_y", None) is None or len(self.last_y) == 0:
            try:
                self.run_current()
            except Exception:
                pass
        if getattr(self, "last_y", None) is None or len(self.last_y) == 0:
            messagebox.showinfo("No sequence", "Please Generate & Plot a sequence first.")
            return

        # 1) Parse inputs
        try:
            dwell_ms = float(self.dwell_var.get())
            dwell_s = max(0.0, dwell_ms / 1000.0)
            compliance = float(self.ilim_var.get())
        except Exception:
            messagebox.showerror("Input error", "Compliance and dwell must be numeric.")
            return

        if not self.device.connected:
            messagebox.showinfo("Not connected", "Connect to the instrument first.")
            return

        # 2) Live plot + status (live window only)
        self.status_var.set("Streaming sequence to instrument...")
        self.last_data = []
        self.live.set_enabled(self.live_var.get() == "On")
        self.live.clear()

        # 运行时只在独立窗口画
        self._ensure_live_window()

        # 根据“Analysis plot”选择决定 x 轴
        sel = (self.analysis_var.get() or "").strip().lower()
        if sel == "i vs time":
            xmode = "time"
        elif sel == "i vs v":
            xmode = "v"
        else:
            xmode = "i vs sample index"  # 保存原文本以便后面复用
            xmode = "index"

        
        self._current_xmode = xmode 
        self._live_axis_applied = False            # <- 新增：标记还没在首帧应用过
   
        # 立刻把 live 窗口坐标轴设为所选模式
        # Set up axes first (clear=True), then attach live target and clear its buffer
        self._config_axes_for_mode(self.live_window.ax, xmode, title=None, clear=True)
        self.live.set_target(self.live_window.ax, self.live_window.canvas)
        self.live.clear()
        self.live_window.canvas.draw_idle()



        # 3) Channel mode (from dropdown; default A only)
        mode = "A only"
        try:
            mode = (self.chan_mode_var.get() or "A only").strip()
        except Exception:
            pass

        # Prepare A/B waveforms (single series -> treat as A)
        yA = getattr(self, "last_yA", None)
        yB = getattr(self, "last_yB", None)
        if yA is None and self.last_y is not None:
            yA = self.last_y

        # If Dual requested but only one series present, synthesize B as inverted A
        if ("Dual" in mode) and yB is None and yA is not None:
            yB = -np.asarray(yA, dtype=float)

        # Keep same length when both exist
        if yA is not None and yB is not None:
            n = min(len(yA), len(yB))
            yA = np.asarray(yA, dtype=float)[:n]
            yB = np.asarray(yB, dtype=float)[:n]

        # 4) Acquisition callback (thread-safe; no Tk calls)
        t0 = time.perf_counter()
        t_prev = t0

        def _cb(rec):
                nonlocal t_prev
                t_now = time.perf_counter()
                dt = t_now - t_prev
                t_prev = t_now

                # 记录绝对运行时间（I vs time 用）
                try:
                        rec["t"] = t_now - t0
                except Exception:
                        pass

                # ---- 统一/补齐键名：I_meas ----
                i_val = rec.get("I_meas")
                if i_val is None:
                        for k in ("I", "current", "curr", "i"):
                                if rec.get(k) is not None:
                                        i_val = rec[k]
                                        break
                rec["I_meas"] = i_val

                # ---- 统一/补齐键名：V_meas ----
                v_val = rec.get("V_meas")
                if v_val is None:
                        for k in ("V", "voltage", "volt", "v"):
                                if rec.get(k) is not None:
                                        v_val = rec[k]
                                        break
                # 还没有的话，用设定值或按 index 从输出序列推断
                if v_val is None:
                        v_val = rec.get("V_set")
                if v_val is None:
                        try:
                                idx_live = int(rec.get("index"))
                                if yA is not None and 0 <= idx_live < len(yA):
                                        v_val = float(yA[idx_live])
                        except Exception:
                                pass
                rec["V_meas"] = v_val

                # 控制台输出（可选）
                try:
                        print(f"[acq] idx={rec.get('index')} dt={dt:.6f}s I={rec.get('I_meas')}")
                except Exception:
                        pass

                # 全量日志
                self.last_data.append(rec)

                # ---- 首帧：再次应用坐标轴文字，防止被 LivePlotService 覆盖 ----
                if not getattr(self, "_live_axis_applied", False):
                        def _apply():
                                try:
                                        self._config_axes_for_mode(
                                                self.live_window.ax,
                                                getattr(self, "_current_xmode", "index"),
                                                title=None,
                                                clear=False  # 不清线，保留 live 线条
                                        )
                                        self.live_window.canvas.draw_idle()
                                finally:
                                        self._live_axis_applied = True

                        try:
                                self.root.after(0, _apply)   # 在 Tk 主线程执行
                        except Exception:
                                pass

                # ---- Live 推点（只往独立窗口；按所选 x 轴）----
                if getattr(self.live, "enabled", False):
                        try:
                                # 计算 X
                                if xmode == "time":
                                        x = rec.get("t") or (time.perf_counter() - t0)
                                elif xmode == "v":
                                        x = rec.get("V_meas")
                                else:  # 'index'
                                        x = rec.get("index")

                                y = rec.get("I_meas")

                                # 仅当数值有效时推送
                                if x is not None and y is not None:
                                        xf = float(x)
                                        yf = float(y)
                                        if np.isfinite(xf) and np.isfinite(yf):
                                                self.live.push(xf, yf)
                        except Exception:
                                pass


        # 5) Worker thread: call controller according to mode
        def _worker():
            import inspect
            err = None
            try:
                # Does output_sequence support "channel"?
                seq_accepts_channel = False
                try:
                    sig = inspect.signature(self.device.output_sequence)
                    seq_accepts_channel = "channel" in sig.parameters
                except Exception:
                    pass

                if mode == "A only":
                    if yA is None:
                        raise RuntimeError("No waveform for channel A.")
                    if seq_accepts_channel:
                        self.device.output_sequence(
                            yA, dwell_s=dwell_s, compliance=compliance, on_point=_cb, channel="A"
                        )
                    else:
                        # Legacy single-channel controller
                        self.device.output_sequence(
                            yA, dwell_s=dwell_s, compliance=compliance, on_point=_cb
                        )

                elif mode == "B only":
                    yB_local = yB if yB is not None else yA
                    if yB_local is None:
                        raise RuntimeError("No waveform for channel B.")
                    if not seq_accepts_channel:
                        raise RuntimeError(
                            "Controller does not support selecting channel='B'. "
                            "Add a 'channel' argument to output_sequence in keithley_controller.py."
                        )
                    self.device.output_sequence(
                        yB_local, dwell_s=dwell_s, compliance=compliance, on_point=_cb, channel="B"
                    )

                elif mode in (
                    "Dual: A = +, B = -",
                    "Dual: A = -, B = +",
                    "Dual: A = +, B = −",
                    "Dual: A = −, B = +",
                ):
                    if yA is None or yB is None:
                        raise RuntimeError("Dual mode requires two waveforms (yA and yB).")

                    # Polarity mapping
                    if "A = -, B = +" in mode or "A = −, B = +" in mode:
                        signA, signB = -1.0, 1.0
                    else:  # "A = +, B = -"
                        signA, signB = 1.0, -1.0
                    yA_local = signA * yA
                    yB_local = signB * yB

                    # Prefer dedicated dual-channel driver
                    if hasattr(self.device, "output_sequence_dual"):
                        self.device.output_sequence_dual(
                            yA_local, yB_local, dwell_s=dwell_s, compliance=compliance, on_point=_cb
                        )
                    else:
                        raise RuntimeError(
                            "Dual-channel output requested but 'output_sequence_dual' is not implemented "
                            "in keithley_controller.py. Implement it or run two synchronized per-channel loops."
                        )
                else:
                    raise ValueError(f"Unknown channel mode: {mode}")

            except Exception as e:
                err = e
            finally:
            
                def _finish():
                    if err:
                        messagebox.showerror("Output error", str(err))
                        self.status_var.set("Instrument: ready (last run failed)")
                        return

                    n = len(self.last_data)
                    self.status_var.set(f"Done. Points streamed: {n}. Output OFF.")

                    try:
                        self.live.set_enabled(False)
                        # 解绑：把 live 的目标切回右侧主面板，防止清空分析图
                        self.live.set_target(self.plot_host.ax, self.plot_host.canvas)
                        self.live.clear()
                    except Exception:
                        pass

                    try:
                        self._plot_analysis_in_window(dwell_s)
                    except Exception as e2:
                        try:
                            print("[finish] analysis plotting failed:", e2)
                        except Exception:
                            pass

                # ★★★ 别忘了真正调用它 ★★★
                try:
                    self.root.after(0, _finish)
                except Exception:
                    pass



        # 6) Launch worker (avoid duplicate runs)
        if self.output_thread and self.output_thread.is_alive():
            messagebox.showinfo("Busy", "Output already running.")
            return
        self.output_thread = threading.Thread(target=_worker, daemon=True)
        self.output_thread.start()



        


    def stop_output(self):
        try:
            if hasattr(self.device, "stop"):
                self.device.stop()
            elif hasattr(self.device, "abort"):
                self.device.abort()
            else:
                # 最差情况：设置一个请求停止的标志（如果你的驱动支持检查它）
                try:
                    setattr(self.device, "_stop_requested", True)
                except Exception:
                    pass
            self.status_var.set("Stopping... (output will turn OFF)")
        except Exception as e:
            messagebox.showerror("Stop error", str(e))


    def save_log_csv(self):
        if not self.last_data:
            messagebox.showinfo("No data", "No measurement log to save.")
            return
        try:
            import pandas as _pd
        except Exception:
            _pd = None

        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV file", "*.csv"), ("All files", "*.*")]
        )
        if not path:
            return
        try:
            if _pd is not None:
                _pd.DataFrame(self.last_data).to_csv(path, index=False)
            else:
                # Minimal fallback CSV writer
                import csv
                keys = ["index", "V_set", "I_meas", "V_meas"]
                with open(path, "w", newline="") as f:
                    w = csv.DictWriter(f, fieldnames=keys)
                    w.writeheader()
                    for row in self.last_data:
                        w.writerow({k: row.get(k, "") for k in keys})
            messagebox.showinfo("Saved", f"Saved {len(self.last_data)} rows to:\n{path}")
        except Exception as e:
            messagebox.showerror("Save error", str(e))

# ---------- Entrypoint ----------
if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)

    def _on_close():
        try:
            app.stop_output()
            app.device.disconnect()
            app.live.stop()
        except Exception:
            pass
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", _on_close)
    root.mainloop()
